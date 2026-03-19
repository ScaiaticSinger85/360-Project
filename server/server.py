from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
import os
import uuid

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EVENTS_FILE = os.path.join(BASE_DIR, "events.xlsx")
EVENTS_SHEET = "Events"

USERS_FILE = os.path.join(BASE_DIR, "users.xlsx")
USERS_SHEET = "Users"


def initialize_events_excel():
    if not os.path.exists(EVENTS_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = EVENTS_SHEET
        sheet.append([
            "Title",
            "Description",
            "Category",
            "Date",
            "Time",
            "Location",
            "Address",
            "Capacity",
            "Image URL",
            "Organizer",
            "Organizer ID",
            "Is Public",
        ])
        workbook.save(EVENTS_FILE)


def initialize_users_excel():
    if not os.path.exists(USERS_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = USERS_SHEET
        sheet.append([
            "User ID",
            "Name",
            "Email",
            "Password Hash",
            "Role",
        ])
        workbook.save(USERS_FILE)


def get_users_sheet():
    initialize_users_excel()
    workbook = load_workbook(USERS_FILE)
    sheet = workbook[USERS_SHEET]
    return workbook, sheet


def get_events_sheet():
    initialize_events_excel()
    workbook = load_workbook(EVENTS_FILE)
    sheet = workbook[EVENTS_SHEET]
    return workbook, sheet


def find_user_by_email(email):
    workbook, sheet = get_users_sheet()
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            user_id, name, saved_email, password_hash, role = row
            if str(saved_email).strip().lower() == email.strip().lower():
                return {
                    "id": str(user_id),
                    "name": str(name),
                    "email": str(saved_email),
                    "password_hash": str(password_hash),
                    "role": str(role),
                }
        return None
    finally:
        workbook.close()


@app.route("/api/auth/signup", methods=["POST"])
def signup():
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                "success": False,
                "message": "No data received."
            }), 400

        name = str(data.get("name", "")).strip()
        email = str(data.get("email", "")).strip().lower()
        password = str(data.get("password", "")).strip()

        if not name or not email or not password:
            return jsonify({
                "success": False,
                "message": "Please fill in all fields."
            }), 400

        if len(name) < 2:
            return jsonify({
                "success": False,
                "message": "Name must be at least 2 characters."
            }), 400

        if "@" not in email or "." not in email:
            return jsonify({
                "success": False,
                "message": "Please enter a valid email address."
            }), 400

        if len(password) < 6:
            return jsonify({
                "success": False,
                "message": "Password must be at least 6 characters."
            }), 400

        existing_user = find_user_by_email(email)
        if existing_user:
            return jsonify({
                "success": False,
                "message": "An account with this email already exists."
            }), 400

        user_id = str(uuid.uuid4())
        password_hash = generate_password_hash(password)
        role = "user"

        workbook, sheet = get_users_sheet()
        try:
            sheet.append([user_id, name, email, password_hash, role])
            workbook.save(USERS_FILE)
        finally:
            workbook.close()

        return jsonify({
            "success": True,
            "message": "Account created successfully.",
            "user": {
                "id": user_id,
                "name": name,
                "email": email,
                "role": role,
            }
        }), 201

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Server error: {str(e)}"
        }), 500


@app.route("/api/auth/signin", methods=["POST"])
def signin():
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                "success": False,
                "message": "No data received."
            }), 400

        email = str(data.get("email", "")).strip().lower()
        password = str(data.get("password", "")).strip()

        if not email or not password:
            return jsonify({
                "success": False,
                "message": "Please fill in all fields."
            }), 400

        user = find_user_by_email(email)

        if not user:
            return jsonify({
                "success": False,
                "message": "No account found with that email."
            }), 401

        if not check_password_hash(user["password_hash"], password):
            return jsonify({
                "success": False,
                "message": "Incorrect password."
            }), 401

        return jsonify({
            "success": True,
            "message": "Signed in successfully.",
            "user": {
                "id": user["id"],
                "name": user["name"],
                "email": user["email"],
                "role": user["role"],
            }
        }), 200

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Server error: {str(e)}"
        }), 500


@app.route("/api/events", methods=["POST"])
def create_event():
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                "success": False,
                "message": "No JSON data received."
            }), 400

        required_fields = [
            "title",
            "description",
            "category",
            "date",
            "time",
            "location",
            "address",
            "capacity",
            "imageUrl",
            "organizer",
            "organizerId",
            "isPublic",
        ]

        missing_fields = []
        for field in required_fields:
            value = data.get(field)
            if value is None or value == "":
                missing_fields.append(field)

        if missing_fields:
            return jsonify({
                "success": False,
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400

        cleaned_data = {
            "title": str(data["title"]).strip(),
            "description": str(data["description"]).strip(),
            "category": str(data["category"]).strip(),
            "date": str(data["date"]).strip(),
            "time": str(data["time"]).strip(),
            "location": str(data["location"]).strip(),
            "address": str(data["address"]).strip(),
            "capacity": int(data["capacity"]),
            "imageUrl": str(data["imageUrl"]).strip(),
            "organizer": str(data["organizer"]).strip(),
            "organizerId": str(data["organizerId"]).strip(),
            "isPublic": bool(data["isPublic"]),
        }

        workbook, sheet = get_events_sheet()
        try:
            sheet.append([
                cleaned_data["title"],
                cleaned_data["description"],
                cleaned_data["category"],
                cleaned_data["date"],
                cleaned_data["time"],
                cleaned_data["location"],
                cleaned_data["address"],
                cleaned_data["capacity"],
                cleaned_data["imageUrl"],
                cleaned_data["organizer"],
                cleaned_data["organizerId"],
                "Yes" if cleaned_data["isPublic"] else "No",
            ])
            workbook.save(EVENTS_FILE)
        finally:
            workbook.close()

        return jsonify({
            "success": True,
            "message": "Event saved successfully.",
            "event": cleaned_data
        }), 201

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Server error: {str(e)}"
        }), 500


if __name__ == "__main__":
    initialize_users_excel()
    initialize_events_excel()
    app.run(host="0.0.0.0", port=4000, debug=False)